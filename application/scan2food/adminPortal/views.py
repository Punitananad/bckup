from django.shortcuts import render, redirect
from .decorator import login_required
from theatre.models import Theatre
from django.contrib.auth.models import Permission
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import datetime, timedelta, time
from theatre.models import Payment as theatre_payment
from theatre.models import Order, FoodItem, OrderRefundRequest, Hall
from django.db.models import Sum
from . models import *
from .form import *
from django.contrib import messages
from django.http import Http404, HttpResponseRedirect, JsonResponse
import os
from django.conf import settings as django_settings
from django.utils.timezone import localtime
import segno
import requests
from requests.auth import HTTPBasicAuth

import pandas as pd
from openpyxl import Workbook, load_workbook

# imports for razorpay settlments
import razorpay
from datetime import timezone as dt_timezone

# webhook
from theatre.update_websocket import update_websocket, send_whatsapp_template

# imports for cahsfree gateways
from cashfree_pg.models.create_order_request import CreateOrderRequest
from cashfree_pg.models.order_create_refund_request import OrderCreateRefundRequest
from cashfree_pg.api_client import Cashfree
from cashfree_pg.models.customer_details import CustomerDetails
from cashfree_pg.models.order_meta import OrderMeta
# ends for cashfree payments ends 

# imports for phonepe gateways
from uuid import uuid4
from phonepe.sdk.pg.payments.v2.standard_checkout_client import StandardCheckoutClient
from phonepe.sdk.pg.payments.v2.models.request.standard_checkout_pay_request import StandardCheckoutPayRequest
from phonepe.sdk.pg.env import Env
from phonepe.sdk.pg.common.models.request.refund_request import RefundRequest
# end for phonepe gateways

# imports for CCAvenue gateways
from Crypto.Cipher import AES
from hashlib import md5
import base64

# imports for PayU gateways
import hashlib

def pad(data):
    length = 16 - (len(data) % 16)
    return data + chr(length) * length

def unpad(data):
    return data[:-ord(data[-1])]

def encrypt(plainText, workingKey):
    iv = bytes([i for i in range(16)])
    encDigest = md5(workingKey.encode('utf-8')).digest()
    plainText = pad(plainText)
    cipher = AES.new(encDigest, AES.MODE_CBC, iv)
    encryptedText = cipher.encrypt(plainText.encode('utf-8'))
    return encryptedText.hex()  # or use base64.b64encode(encryptedText).decode()

def decrypt(cipherText, workingKey):
    iv = bytes([i for i in range(16)])
    decDigest = md5(workingKey.encode('utf-8')).digest()
    encryptedBytes = bytes.fromhex(cipherText)
    cipher = AES.new(decDigest, AES.MODE_CBC, iv)
    decryptedText = cipher.decrypt(encryptedBytes).decode('utf-8')
    return unpad(decryptedText)

# ends for CCAvenue gateways


@login_required
# Create your views here.
def index(request):
    return render(request, 'adminPortal/index.html')

@login_required
def all_theatre(request):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        all_theatre = Theatre.objects.all()
        paginator = Paginator(all_theatre, 20)  # Show 20 Theatre Per Page.

        page_number = request.GET.get('page')
        if page_number == None:
            page_number = 1
        all_theatre = paginator.get_page(page_number)
        return render(request, 'adminPortal/all-theatre.html', {'all_theatre': all_theatre})
    else:
        return HttpResponse('Permission Denied')

@login_required
def theatre_detail(request, pk):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        try:
            theatre = Theatre.objects.get(pk=pk)
            document_form = DocumentForm()
            logo_form = ImageUploadform()
            try:
                gst_form = GSTDetailsForm(instance=theatre.gstdetails)
                gst_details = theatre.gstdetails
            except:
                gst_form = GSTDetailsForm()
                gst_details = None

            try:
                bank_form = BankDetailsForm(instance=theatre.bankdetails)
                bank_details = theatre.bankdetails
            except:
                bank_form = BankDetailsForm()
                bank_details = None

            return render(request, 'adminPortal/theatre-detail.html', {'theatre': theatre,'bank_form': bank_form, 'document_form': document_form, 'logo_form': logo_form, 'gst_form': gst_form, 'gst_details': gst_details, 'bank_details': bank_details})
        except:
            raise Http404('page not found')
    else:
        return HttpResponse('Permission Denied')

@login_required
def theatre_profile(request):
    return render(request, 'adminPortal/theatre-profile.html')

@login_required
def update_gst_detail(request,pk):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        try:
            theatre = Theatre.objects.get(pk=pk)
            if request.method == 'POST':
                try:
                    gst_instance = theatre.gstdetails
                except:
                    gst_instance = GSTDetails.objects.create(theatre=theatre)

                form = GSTDetailsForm(request.POST, instance=gst_instance)
                if form.is_valid():
                    form.save()
                    messages.success(request, 'GST Details Updated Successfully')
                else:
                    messages.error(request, 'Not able to update the GST Details')

                return redirect('admin-portal:theatre-detail', pk)

        except:
            raise Http404('page not found')
    else:
        return HttpResponse('Permission Denied')

@login_required
def update_bank_detail(request, pk):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        try:
            theatre = Theatre.objects.get(pk=pk)
            if request.method == 'POST':
                try:
                    bank_instance = theatre.bankdetails
                except:
                    bank_instance = bankDetails.objects.create(theatre=theatre)

                form = BankDetailsForm(request.POST, instance=bank_instance)
                if form.is_valid():
                    form.save()
                    messages.success(request, 'Bank Details Updated Successfully')
                else:
                    messages.error(request, 'Not able to update the Bank Details')

                return redirect('admin-portal:theatre-detail', pk)

        except:
            raise Http404('page not found')
    else:
        return HttpResponse('Permission Denied')

@login_required
def update_logo(request, pk):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        try:
            theatre = Theatre.objects.get(pk=pk)
            if request.method == 'POST':
                form = ImageUploadform(request.POST, request.FILES)
                if form.is_valid():
                    file = form.cleaned_data['file']  # Get the file from form data
                    file_path = os.path.join(django_settings.MEDIA_ROOT, 'theatre_logo', file.name)

                    if os.path.exists(file_path):
                        folder_path = os.path.join(django_settings.MEDIA_ROOT, 'theatre_logo')
                        # count the number of files and add number in starting
                        add_on = len([f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))])
                        file_path = os.path.join(django_settings.MEDIA_ROOT, 'theatre_logo', f"{add_on}{file.name}")
                        file_path = file_path.replace(" ", "-")
                    
                    # Save the file to the media directory
                    with open(file_path, 'wb+') as destination:
                        for chunk in file.chunks():
                            destination.write(chunk)

                    theatre.detail.logo = f"theatre_logo/{file.name}"
                    theatre.detail.save()
                    messages.success(request, 'Logo Updated Successfully')
                else:
                    messages.error(request, 'Not able to upload the file')

                return redirect('admin-portal:theatre-detail', pk)

        except:
            raise Http404('page not found')
    else:
        return HttpResponse('Permission Denied')


@login_required
def upload_document(request, pk):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        if request.method == 'POST':
            form = DocumentForm(request.POST, request.FILES)

            if form.is_valid():
                document_instance = form.save(commit=False)
                document_instance.theatre = Theatre.objects.get(pk=pk)

                # Ensure the directory exists
                folder_path = os.path.join(django_settings.MEDIA_ROOT, 'documents', str(pk))
                os.makedirs(folder_path, exist_ok=True)

                document_instance.save()
                messages.success(request, 'Document Uploaded Successfully')
            
            else:
                messages.error(request, 'Not able to upload the file')

            return redirect('admin-portal:theatre-detail', pk)
    else:
        return HttpResponse('Permission Denied')

@login_required
def delete_document(request, pk):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        try:
            document = Documents.objects.get(pk=pk)
            theatre = document.theatre
            document_location = document.document.path
            # Delete the file from the filesystem
            if os.path.exists(document_location):
                os.remove(document_location)

            document.delete()
            messages.success(request, 'Document Deleted Successfully')
            return redirect('admin-portal:theatre-detail', theatre.pk)
        except:
            raise Http404('page not found')
    else:
        return HttpResponse('Permission Denied')

@login_required
def new_payout_payments(request, pk):
    permission = Permission.objects.get(codename="add_payoutlogs")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        try:
            theatre = Theatre.objects.get(pk=pk)
            pending_payments = theatre_payment.objects.filter(is_settled=False, status='Success', payment_method='Gateway', order__seat__row__hall__theatre=theatre).order_by('time')

            pending_amount = 0

            first_payment = pending_payments.first()
            
            if first_payment == None:
                first_payment_time = timezone.now()

            else:
                first_payment_time = timezone.localtime(first_payment.time)
                pending_amount = sum(payment.order.full_payment() - payment.settlement for payment in pending_payments)
                

            starting_hour = first_payment_time.hour

            # CHECK THE HOUR IS LESS THEN 6
            if starting_hour <= 6:
                # GET ONE DAY BACK
                previous_day = first_payment_time - timedelta(days=1)
                
                # GET 8:00 TO NEXT 6:00 TIME
                start_time = previous_day.replace(hour=6, minute=0, second=0, microsecond=0)

                end_time = first_payment_time.replace(hour=6, minute=0, second=0, microsecond=0)


            else:
                # GET CURRENT DAY COMPLETE DATA
                
                # GET NEXT DAY
                next_day = first_payment_time + timedelta(days=1)
                # GET 8:00 TO NEXT 6:00 TIME
                start_time = first_payment_time.replace(hour=6, minute=0, second=0, microsecond=0)
                
                end_time = next_day.replace(hour=6, minute=0, second=0, microsecond=0)

            # GET PAYMENTS BETWEEN THIS TIME
            payments = pending_payments.filter(time__gt=start_time, time__lt=end_time)
            current_balance = sum(payment.order.order_amount() for payment in payments)

            if request.method == 'POST':
                # GENERATE A NEW PAYOUT
                if payments.count() == 0:
                    messages.error(request, f'There is No Payment to generate Payout for date:- {start_time.strftime("%d-%b-%Y")}')

                else:
                    payout = PayOutLogs.objects.create(start_time=start_time, end_time=end_time, user=request.user, amount=current_balance, theatre=theatre)
                    for payment in payments:
                        payment.payout = payout
                        payment.is_settled = True
                        payment.save()

                    messages.success(request, f'Payout is generated of date {start_time.strftime("%d-%b-%Y")}')

                return redirect('admin-portal:new-payout-payments', theatre.pk)
            
            context = {
                "payments": payments,
                "current_balance": current_balance,
                "theatre": theatre,
                "start_time": start_time,
                "end_time": end_time,
                "pending_amount": pending_amount,
            }
            return render(request, 'adminPortal/new-payouts-payment.html', context)
        
        except:
            raise Http404('page not found')
        
    else:
        return HttpResponse('Permission Denied')

@login_required
def all_payouts(request):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        date_range = request.GET.get('daterange', "")
        if date_range == "":
            start_date = timezone.now().date()
            end_date = timezone.now().date()
        else:
            try:
                start_date = date_range.split(" - ")[0]
                end_date = date_range.split(" - ")[1]

                start_date = datetime.strptime(start_date, "%d/%b/%Y").date()
                end_date = datetime.strptime(end_date, "%d/%b/%Y").date()
            except:
                start_date = timezone.now().date()
                end_date = timezone.now().date()
        
        all_payouts = PayOutLogs.objects.all().order_by('-start_time')
        paginator = Paginator(all_payouts, 10)

        page_number = request.GET.get('page')
        if page_number == None:
            page_number = 1
        
        all_payouts = paginator.get_page(page_number)

        context = {
            "all_payouts": all_payouts,
            "start_date": start_date,
            "end_date": end_date
        }

        return render(request, 'adminPortal/all-payouts.html', context)

    else:
        return HttpResponse('Permission Denied')
    
@login_required
def payout_payments(request, pk):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        try:
            payout = PayOutLogs.objects.get(pk=pk)
            payments = payout.payment_set.order_by('-time')
            return render(request, 'adminPortal/payout-payments.html', {"payout": payout, "payments": payments})

        except:
            raise Http404('page not found')
        
    else:
        return HttpResponse('Permission Denied')

@login_required
def update_payout_settlement(request, pk):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        try:
            payout = PayOutLogs.objects.get(pk=pk)
            return_data = {'status': ''}
            if payout.is_settled:
                payout.is_settled = False
                return_data['status'] = False
                
            else:
                payout.is_settled = True
                return_data['status'] = True
            
            payout.save()
            return JsonResponse(return_data)
        
        except:
            return Http404('page not found')
        
    else:
        return HttpResponse('Permission Denied')
    
@login_required
def settings(request):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        return render(request, 'adminPortal/settings.html')
    else:
        return HttpResponse('Permission Denied')
    
@login_required
def gateways(request):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        if request.method == 'POST':
            form = PaymentGatewayForm(request.POST)
            if form.is_valid():
                form.save()
                
                messages.success(request, 'Payment Gateway Added Successfully')

            return redirect('admin-portal:gateways')
        
        form = PaymentGatewayForm()
        all_gateways = PaymentGateway.objects.all()
        context = {
            "form": form,
            "all_gateways": all_gateways
        }
        return render(request, 'adminPortal/gateways.html', context)
    else:
        return HttpResponse('Permission Denied')

@login_required
def activate_gateway(request, pk):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        try:
            # DEACTIVATE ALL GATEWAYS
            PaymentGateway.objects.all().update(is_active=False)
            # ACTIVATE THE SELECTED GATEWAY
            gateway = PaymentGateway.objects.get(pk=pk)
            gateway.is_active = not gateway.is_active
            gateway.save()
            messages.success(request, f'{gateway.name} Gateway is Activated Successfully')
            return redirect('admin-portal:gateways')
        except:
            return Http404('page not found')
    else:
        return HttpResponse('Permission Denied')
    

@login_required
def all_orders_new_page(request):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        date_range = request.GET.get('daterange', "")
        order_status = request.GET.get('order-status')
        selected_theatre = request.GET.get('selected-theatre', "")

        deactivate_theatres = Theatre.objects.filter(detail__scaning_service=False).count()

        current_time = localtime(timezone.now())
        current_time = current_time.replace(tzinfo=None)

        if date_range == "":
            start_date = timezone.now().date()
            end_date = timezone.now().date()

            if current_time.hour < 6:
                start_date = start_date - timedelta(days=1)
            else:
                end_date = end_date + timedelta(days=1)

        else:
            try:
                start_date = date_range.split(" - ")[0]
                end_date = date_range.split(" - ")[1]

                start_date = datetime.strptime(start_date, "%d/%b/%Y").date()
                end_date = datetime.strptime(end_date, "%d/%b/%Y").date()
            except:
                start_date = localtime(timezone.now()).date()
                end_date = localtime(timezone.now()).date()

                if current_time.hour < 6:
                    start_date = start_date - timedelta(days=1)
                else:
                    end_date = end_date + timedelta(days=1)
        
        tiem_obj = time(hour=6, minute=0, second=0, microsecond=0)
        start_time = datetime.combine(start_date, tiem_obj)
        end_time = datetime.combine(end_date, tiem_obj)
        
        if selected_theatre == "":
            all_orders = Order.objects.filter(start_time__range=(start_time, end_time))
        else:
            all_orders = Order.objects.filter(start_time__range=(start_time, end_time), seat__row__hall__theatre__pk=selected_theatre)
            
        total_amount = theatre_payment.objects.filter(order__in=all_orders, status="Success").aggregate(total=Sum("amount"))["total"] or 0

        if order_status == None or order_status == "":
            order_status = "Success"

        if order_status != "All":
            all_orders = all_orders.filter(payment__status=order_status)

        all_orders = all_orders.order_by('-start_time')

        theatre_amount = 0
        commission = 0
        for order in all_orders:
            theatre_amount += order.full_payment()
            commission += order.payment.settlement
        
        total_amount = round(total_amount, 2)

        net_profit = total_amount - theatre_amount + commission
        net_profit = round(net_profit, 2)
        ordr_count = all_orders.count()

        paginator = Paginator(all_orders, 25)  # Show 25 items per page.

        page_number = request.GET.get('page')
        if page_number == None:
            page_number = 1
        orders = paginator.get_page(page_number)

        context = {
            'orders': orders,
            "start_date": start_date,
            "end_date": end_date,
            "order_status": order_status,
            "net_profit": net_profit,
            "ordr_count": ordr_count,
            "theatre_amount": theatre_amount,
            "total_amount": total_amount,
            "selected_theatre": selected_theatre,
            "deactivate_theatres": deactivate_theatres
            }

        return render(request, 'adminPortal/all-orders-new.html', context)
    
    else:
        return HttpResponse('Permission Denied')

@login_required
def all_orders(request):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        deactivate_theatres = Theatre.objects.filter(detail__scaning_service=False).count()
        context = {
            "deactivate_theatres": deactivate_theatres
        }
        return render(request, 'adminPortal/all-orders.html', context)
    else:
        return HttpResponse('Permission Denied')

@login_required
def live_orders(request):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        return render(request, 'adminPortal/live-orders.html')
    
    else:
        return HttpResponse('Permission Denied')

@login_required
def order_profile(request, pk):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        try:
            order = Order.objects.get(pk=pk)
            theatre = order.seat.row.hall.theatre
            return render(request, 'adminPortal/order-profile.html', {"order_id": order.id, 'theatre': theatre})
        except:
            raise Http404('page not found')
    else:
        return HttpResponse('Permission Denied')

@login_required
def refund_order(request, pk):
    permission = Permission.objects.get(codename="delete_order")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        order = Order.objects.get(pk=pk)
        reason = request.POST.get('Refunded By Admin')
        payment = order.payment
        if payment.payout == None:
            if payment.payment_method == 'Gateway':
                
                if payment.status == 'Refunded':
                    messages.error(request, 'Payment has already been settled and cannot be refunded.')

                else:
                    refund_generated = False
                    if payment.gateway == "split_razorpay":
                        # CODE IS WORKING AND COMPLETED FOR SPLIT RAZORPAY
                        pay_id = payment.splitrazorpaypayment.razorpay_payment_id
                        amount = payment.amount * 100
                        refund_url = f'https://api.razorpay.com/v1/payments/{pay_id}/refund'
                        payload = {
                            'amount': amount,
                            "reverse_all": True
                        }
                        gateway_detail = PaymentGateway.objects.get(name='split_razorpay')
                        response = requests.post(
                            refund_url,
                            auth=HTTPBasicAuth(gateway_detail.gateway_key, gateway_detail.gateway_secret),
                            data=payload
                        )

                        if response.status_code == 200:
                            seat = order.seat
                            seat_last_order = seat.order_set.filter(payment__status='Success').last()
                            if seat_last_order == order:
                                seat.is_vacent = True
                                # run Channel Code
                                # Update the websocket
                                update_websocket(
                                    theatre_id=seat.row.hall.theatre.id,
                                    seat_id=seat.id,
                                    is_vacent=seat.is_vacent,
                                    payment_panding=False,
                                    seat_name=f"{seat.row.hall.name} | {seat.name}",
                                    message='Refund is Generated and It Will Transfer within 5-7 Working Days !',
                                    customer_phone=order.payment.phone_number,
                                    customer_message="✅ Refund Confirmed! %0A%0AYour refund has been successfully processed. 💸%0A%0AThe amount will be credited to your original payment method within 5–7 working days. 🏦%0A%0AThank you for your patience and for choosing Scan2Food. 🙏%0A%0AFor any queries, feel free to contact us at support@scan2food.com 📧%0A%0A— Team Scan2Food 🍿",
                                    theatre_name=seat.row.hall.theatre.name,
                                    msg_typ="refund-conformation",
                                    order_id=payment.order.pk
                                )

                                seat.save()

                            payment.status = 'Refunded'
                            order.delivery_time = localtime(timezone.now())
                            payment.save()
                            order.save()
                            refund_generated = True
                            messages.success(request, 'Refund is Generated and It Will Transfer within 5-7 Working Days !')
                        else:
                            messages.error(request, 'No payment was collected for this Order')

                    elif payment.gateway == 'Razorpay':
                        # CODE IS WORKING AND COMPLETED FOR RAZORPAY
                        pay_id = payment.razorpaypayment.razorpay_payment_id
                        amount = payment.amount * 100
                        refund_url = f'https://api.razorpay.com/v1/payments/{pay_id}/refund'
                        payload = {
                            'amount': amount
                        }
                        gateway_detail = PaymentGateway.objects.get(name='Razorpay')
                        response = requests.post(
                            refund_url,
                            auth=HTTPBasicAuth(gateway_detail.gateway_key, gateway_detail.gateway_secret),
                            data=payload
                        )

                        if response.status_code == 200:
                            seat = order.seat
                            seat_last_order = seat.order_set.filter(payment__status='Success').last()
                            if seat_last_order == order:
                                seat.is_vacent = True
                                # run Channel Code
                                # Update the websocket
                                update_websocket(
                                    theatre_id=seat.row.hall.theatre.id,
                                    seat_id=seat.id,
                                    is_vacent=seat.is_vacent,
                                    payment_panding=False,
                                    seat_name=f"{seat.row.hall.name} | {seat.name}",
                                    message='Refund is Generated and It Will Transfer within 5-7 Working Days !',
                                    customer_phone=order.payment.phone_number,
                                    customer_message="✅ Refund Confirmed! %0A%0AYour refund has been successfully processed. 💸%0A%0AThe amount will be credited to your original payment method within 5–7 working days. 🏦%0A%0AThank you for your patience and for choosing Scan2Food. 🙏%0A%0AFor any queries, feel free to contact us at support@scan2food.com 📧%0A%0A— Team Scan2Food 🍿",
                                    theatre_name=seat.row.hall.theatre.name,
                                    msg_typ="refund-conformation",
                                    order_id=payment.order.pk
                                )

                                seat.save()

                            payment.status = 'Refunded'
                            order.delivery_time = localtime(timezone.now())
                            payment.save()
                            order.save()
                            refund_generated = True
                            messages.success(request, 'Refund is Generated and It Will Transfer within 5-7 Working Days !')
                        else:
                            messages.error(request, 'No payment was collected for this Order')

                    elif payment.gateway == 'Cashfree':
                        # CODE IS WORKING AND COMPLETED FOR CASHFREE
                        gateway_detail = PaymentGateway.objects.get(name='Cashfree')

                        Cashfree.XClientId = gateway_detail.gateway_key
                        Cashfree.XClientSecret = gateway_detail.gateway_secret
                        Cashfree.XEnvironment = Cashfree.PRODUCTION # SANDBOX or PRODUCTION
                        x_api_version = "2023-08-01"
                        order_id = payment.cashfreepayment.cashfree_order_id
                        try:
                            # Create Refund Order
                            refundRequest = OrderCreateRefundRequest(
                                refund_amount=payment.amount,
                                refund_id=f"refund_p_id_{payment.pk}"
                            )

                            # Refund Request
                            api_response = Cashfree().PGOrderCreateRefund(x_api_version, order_id, refundRequest)
                            
                            if api_response.status_code == '200' or 200:
                                seat = order.seat
                                seat_last_order = seat.order_set.filter(payment__status='Success').last()

                                if seat_last_order == order:
                                    seat.is_vacent = True
                                    # run Channel Code
                                    # Update the websocket
                                    # # update websocket
                                    update_websocket(
                                        theatre_id=seat.row.hall.theatre.id,
                                        seat_id=seat.id,
                                        is_vacent=seat.is_vacent,
                                        payment_panding=False,
                                        seat_name=f"{seat.row.hall.name} | {seat.name}",
                                        message='Refund is Generated and It Will Transfer within 5-7 Working Days !',
                                        customer_phone=order.payment.phone_number,
                                        customer_message="✅ Refund Confirmed! %0A%0AYour refund has been successfully processed. 💸%0A%0AThe amount will be credited to your original payment method within 5–7 working days. 🏦%0A%0AThank you for your patience and for choosing Scan2Food. 🙏%0A%0AFor any queries, feel free to contact us at support@scan2food.com 📧%0A%0A— Team Scan2Food 🍿",
                                        theatre_name=seat.row.hall.theatre.name,
                                        msg_typ="refund-conformation",
                                        order_id=payment.order.pk
                                    )

                                    seat.save()
                                    
                                payment.status = 'Refunded'
                                order.delivery_time = localtime(timezone.now())
                                payment.save()
                                order.save()
                                refund_generated = True
                                messages.success(request, 'Refund is Generated and It Will Transfer within 5-7 Working Days !')

                            else:
                                pass

                        except:
                            messages.error(request, 'No payment was collected for this Order or Refund is Already Generated it may be refunded in 2-3 days in Account')

                    elif payment.gateway == 'Phonepe':
                        # CODE IS PENDING FOR PHONEPE
                        selected_gateway = PaymentGateway.objects.filter(name="Phonepe").first()
                        client_id = selected_gateway.gateway_key
                        client_secret = selected_gateway.gateway_secret
                        client_version = 1  
                        env=Env.PRODUCTION
                        client = StandardCheckoutClient.get_instance(client_id=client_id,
                                                                            client_secret=client_secret,
                                                                            client_version=client_version,
                                                                            env=env)

                        # check about the payment status

                        unique_order_id = payment.phonepepayment.uu_id
                        order_response = client.get_order_status(merchant_order_id=unique_order_id)
                        order_status = order_response.state

                        if order_status == "COMPLETED":
                            try:
                                # PROSCEED IF PAYMENT WAS COMPLETED...
                                unique_merchant_refund_id = f"refund-{unique_order_id}"
                                # unique_merchant_refund_id = str(uuid4())
                                original_merchant_order_id = unique_order_id
                                amount = int(payment.amount * 100)
                                
                                refund_request = RefundRequest.build_refund_request(merchant_refund_id=unique_merchant_refund_id,
                                                                                    original_merchant_order_id=original_merchant_order_id,
                                                                                    amount=amount)
                                refund_response = client.refund(refund_request=refund_request)

                                refund_id = refund_response.refund_id
                                if refund_id:
                                    seat = payment.order.seat
                                    seat_last_order = seat.order_set.filter(payment__status='Success').last()
                                    if seat_last_order == order:
                                        seat.is_vacent = True
                                        # run Channel Code
                                        # Update the websocket
                                        update_websocket(
                                            theatre_id=seat.row.hall.theatre.id,
                                            seat_id=seat.id,
                                            is_vacent=seat.is_vacent,
                                            payment_panding=False,
                                            seat_name=f"{seat.row.hall.name} | {seat.name}",
                                            message='Refund is Generated and It Will Transfer within 5-7 Working Days !',
                                            customer_phone=order.payment.phone_number,
                                            customer_message="✅ Refund Confirmed! %0A%0AYour refund has been successfully processed. 💸%0A%0AThe amount will be credited to your original payment method within 5–7 working days. 🏦%0A%0AThank you for your patience and for choosing Scan2Food. 🙏%0A%0AFor any queries, feel free to contact us at support@scan2food.com 📧%0A%0A— Team Scan2Food 🍿",
                                            theatre_name=seat.row.hall.theatre.name,
                                            msg_typ="refund-conformation",
                                            order_id=payment.order.pk
                                        )

                                        seat.save()

                                        order.delivery_time = localtime(timezone.now())
                                        order.save()

                                    payment.status = 'Refunded'
                                    payment.save()
                                    phonepe_payment = payment.phonepepayment
                                    phonepe_payment.refund_id = refund_id
                                    phonepe_payment.save()
                                    refund_generated = True
                                    messages.success(request, 'Refund is Generated and It Will Transfer within 5-7 Working Days !')

                            except Exception as e:
                                messages.error(request, "This Payment May be already Refunded or No Payment was Collected for this Order !")


                    elif payment.gateway == 'CCAvenue':
                        # CODE IS PENDING FOR CCAvenue
                        pass

                    elif payment.gateway == 'PayU':
                        selected_gateway = PaymentGateway.objects.get(name='PayU')
                        gateway_key = selected_gateway.gateway_key
                        order_id = payment.payupayment.uu_id
                        gateway_salt = selected_gateway.gateway_salt
                        
                        # check whether the payment is successfull or not if the payment is successfull then only prosceed the work...
                        url = "https://info.payu.in/merchant/postservice.php?form=2"
                        
                        # hash to check the payment status
                        hash_string = f"{gateway_key}|verify_payment|{order_id}|{gateway_salt}"
                        hashh = hashlib.sha512(hash_string.encode('utf-8')).hexdigest()

                        payload = {
                            "key": gateway_key,
                            "command": "verify_payment",
                            "var1": order_id,
                            "hash": hashh
                        }

                        response = requests.post(url, data=payload)
                        response_data = response.json()

                        status = response_data['transaction_details'][order_id]['status']

                        if status == 'success':
                            # Payment is success now try to refund the amount
                            command = "cancel_refund_transaction"
                            mihpayid = response_data['transaction_details'][order_id]['mihpayid']  # ✅ MUST be PayU's transaction ID
                            amount = payment.amount

                            # Hash format for refund
                            # correct hash format: key|command|var1|salt
                            hash_string = f"{gateway_key}|{command}|{mihpayid}|{gateway_salt}"
                            hashh = hashlib.sha512(hash_string.encode("utf-8")).hexdigest()
                            url = "https://info.payu.in/merchant/postservice?form=2"

                            payload = {
                                "key": gateway_key,
                                "command": command,
                                "var1": mihpayid,     # ✅ Must be mihpayid, NOT order_id
                                "var2": order_id,
                                "var3": amount,
                                "var5": "https://www.scan2food.com/payu-webhook-url",
                                "hash": hashh
                            }

                            response = requests.post(url, data=payload)
                            response_data = response.json()
                            if response_data['status'] == 1:
                                refund_generated = True
                                messages.success(request, response_data['msg'])
                                seat = order.seat
                                seat_last_order = seat.order_set.filter(payment__status='Success').last()
                                if seat_last_order == order:
                                    seat.is_vacent = True
                                    seat.save()


                                update_websocket(
                                    theatre_id=seat.row.hall.theatre.id,
                                    seat_id=seat.id,
                                    is_vacent=seat.is_vacent,
                                    payment_panding=False,
                                    seat_name=f"{seat.row.hall.name} | {seat.name}",
                                    message='Refund is Generated and It Will Transfer within 5-7 Working Days !',
                                    customer_phone=order.payment.phone_number,
                                    customer_message="✅ Refund Confirmed! %0A%0AYour refund has been successfully processed. 💸%0A%0AThe amount will be credited to your original payment method within 5–7 working days. 🏦%0A%0AThank you for your patience and for choosing Scan2Food. 🙏%0A%0AFor any queries, feel free to contact us at support@scan2food.com 📧%0A%0A— Team Scan2Food 🍿",
                                    theatre_name=seat.row.hall.theatre.name,
                                    msg_typ="refund-conformation",
                                    order_id=payment.order.pk
                                )

                                payment.status = 'Refunded'
                                order.delivery_time = localtime(timezone.now())
                                payment.save()
                                order.save()

                                # refund is initiated in queue and will be refunded in some time...
                            else:
                                if 'Refund FAILURE' in response_data['msg']:
                                    payment.status = 'Refunded'
                                    order.delivery_time = localtime(timezone.now())
                                    payment.save()
                                    order.save()
                                messages.error(request, response_data['msg'])

                        else:
                            # Payment is Not Completed Yet...
                            messages.error(request, 'Payment is not Completed Yet. So Not able to Refund the amount !')
                            
                    payment.reason = reason
                    payment.save() 
                    
                    if refund_generated == True:
                        send_whatsapp_template(phone_number=payment.phone_number, msg_typ='refund-conformation')

                return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))
            else:
                messages.error(request, 'payment was collected via Cash so no refund is possible')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))
        else:
            messages.error(request, 'Payment is Already Sorted and Send to the Theatre Owner !')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    else:
        return HttpResponse('Permission Denied')
    
@login_required
def delete_order(request, pk):
    permission = Permission.objects.get(codename="delete_order")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
    
        try:
            order = Order.objects.get(pk=pk)
            if request.user.groups.first().name == "service_provider":
                order.delete()
                messages.success(request, f'Payment of Order: {pk} has been Deleted Successfully ! ')

            else:
                messages.error(request, "You don't have permission to delete this Order")
                return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))
        except:
            raise Http404('page not found ...')
        
        return redirect('admin-portal:all-orders')
            

@login_required
def upload_food_image(request):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        if request.method == 'POST':
            form = ImageUploadform(request.POST, request.FILES)
            if form.is_valid():
                file = form.cleaned_data['file']  # Get the file from form data

                file_path = os.path.join(django_settings.MEDIA_ROOT, 'food_images', file.name)
                if os.path.exists(file_path):
                    folder_path = os.path.join(django_settings.MEDIA_ROOT, 'food_images')

                    # count the number of files and add number in starting
                    add_on = len([f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))])
                    file_path = os.path.join(django_settings.MEDIA_ROOT, 'food_images', f"{add_on}{file.name}")
                
                # Save the file to the media directory
                with open(file_path, 'wb+') as destination:
                    for chunk in file.chunks():
                        destination.write(chunk)

                messages.success(request, 'image uploaded successfully...')
            else:
                messages.error(request, 'Not able to upload the fiel')

            return redirect('admin-portal:upload-food-image')
        form = ImageUploadform()
        return render(request, 'adminPortal/file-form.html', {'form': form})
        
    else:
        return HttpResponse('Permission Denied')

@login_required
def item_approved_list(request):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        page_number = request.GET.get("page")
        if page_number == None:
            page_number = 1
        all_items = FoodItem.objects.all().order_by('is_approved')
        paginator = Paginator(all_items, 25)
        return render(request, 'adminPortal/item-approved-list.html', {'all_items': paginator.get_page(page_number)})
    
    else:
        raise HttpResponse('Permission Denied')

@login_required
def approve_food_item(request, pk):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        try:
            item = FoodItem.objects.get(pk=pk)
            if item.is_approved:
                item.is_approved = False
            else:
                item.is_approved = True
            
            item.save()

            return JsonResponse({'status': item.is_approved, 'item-name': item.name, 'theatre': item.catogary.theatre.name})
        except:
            raise Http404('error')
    else:
        return HttpResponse('Permission Denied')
    
@login_required
def user_aggriment(request, pk):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        theatre = Theatre.objects.get(pk=pk)
        current_date = localtime(timezone.now())
        data = {'theatre': theatre, "current_date": current_date}
        return render(request, 'adminPortal/user-aggriment.html', data,)
    else:
        return HttpResponse("Permission Denied")

@login_required
def download_report(request, pk):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        try:
            payout = PayOutLogs.objects.get(pk=pk)
            payments = payout.payment_set.all()
            
            all_orders = Order.objects.filter(payment__in=payments).order_by('-start_time')
            all_orders = all_orders.order_by('-start_time')

            data = []
            i = 0
            total_payment = 0
            for order in all_orders:
                theatre = order.seat.row.hall.theatre
                i += 1
                item_list = ""
                for item in order.items.all():
                    item_list += f"{item.name} * {item.quantity}, "
                total_payment += order.full_payment()
                append_data = [
                    f"{i}",
                    f"{order.seat.row.hall.name} ({order.seat.name})",
                    order.pk,
                    item_list,
                    order.full_payment(),
                    localtime(order.start_time).strftime("%d-%b-%Y %I:%M %p"),
                    localtime(order.delivery_time).strftime("%d-%b-%Y %I:%M %p"),
                    ]

                data.append(append_data)
            
            file_path = os.path.join(django_settings.MEDIA_ROOT, 'report.xlsx')
            if os.path.exists(file_path):
                os.remove(file_path)
            
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'Theatre Name '
            ws['C1'] = theatre.name

            ws['A2'] = 'Duration'
            st_date = payout.start_time.strftime("%d-%b-%Y")
            ed_dt = payout.end_time.strftime("%d-%b-%Y")
            ws['C2'] = f"{st_date} 6:00 AM - {ed_dt} 6:00 AM"

            ws['A3'] = 'Total Amount'
            ws['C3'] = total_payment


            ws.append([])
            headers = ["S No.", "Seat", "Order Id", "Item List", "Amount", "Order Initiated Time", "Delivery Time"]
            ws.append(headers)

            # Append Data
            for row in data:
                ws.append(row)
            
            wb.save(file_path)
            # ✅ Return file as response
            with open(file_path, "rb") as excel:
                response = HttpResponse(excel.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                response["Content-Disposition"] = f'attachment; filename="report.xlsx"'

            # Clean up temp file
            os.remove(file_path)
            return response
        except:
            return HttpResponse('There is some problem')
    else:
        return HttpResponse('permission Denied')

@login_required
def download_bulk_report(request):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        
        date_range = request.GET.get('daterange', "")
        if date_range == "":
            start_date = localtime(timezone.now()).date()
            end_date = localtime(timezone.now()).date()
        else:
            try:
                start_date = date_range.split(" - ")[0]
                end_date = date_range.split(" - ")[1]

                start_date = datetime.strptime(start_date, "%d/%b/%Y").date()
                end_date = datetime.strptime(end_date, "%d/%b/%Y").date()
            except:
                start_date = localtime(timezone.now()).date()
                end_date = localtime(timezone.now()).date()

            start_time = datetime.combine(start_date, time(6,0))

            end_time = datetime.combine(end_date, time(6,0))

            all_theatres = Theatre.objects.all()
            pending_payments = theatre_payment.objects.filter(is_settled=False, status='Success', payment_method='Gateway', time__gt=start_time, time__lt=end_time)

            #  create the payment for excel file
            data = []
            for theatre in all_theatres:
                theatre_payments = pending_payments.filter(order__seat__row__hall__theatre=theatre)
                try:
                    bank_details = theatre.bankdetails
                    current_balance = sum(payment.order.full_payment() - payment.settlement for payment in theatre_payments)
                    append_data = [
                        bank_details.account_name,
                        bank_details.account_number,
                        bank_details.ifsc_code,
                        current_balance,
                        theatre.name
                    ]
                    data.append(append_data)

                except:
                    pass

            file_path = os.path.join(django_settings.MEDIA_ROOT, 'bulk_payout.xlsx')
            if os.path.exists(file_path):
                os.remove(file_path)
            
            wb = Workbook()
            ws = wb.active

            headers = ["Beneficiary Name (Mandatory) Full name of the customer - eg: Bruce Wayne"
                       , "Beneficiary Account number (Mandatory) Beneficiary Account number to which the money should be transferred",
                       "IFSC code (Mandatory) IFSC code of beneficary's bank. eg:KKBK0000958",
                       "Amount (Mandatory) Amount that needs to be transfered. Eg: 100.00",
                       "Description / Purpose (Optional) For Internal Reference eg: For salary",
                       ]
            
            ws.append(headers)

            f_name = f'{start_date.strftime("%d-%b-%Y")}:6:00-am|{end_date.strftime("%d-%b-%Y")}:6:00-am.xlsx'
            # Append Data
            for row in data:
                ws.append(row)
            
            wb.save(file_path)
            # ✅ Return file as response
            with open(file_path, "rb") as excel:
                response = HttpResponse(excel.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                response["Content-Disposition"] = f'attachment; filename={f_name}'

            # Clean up temp file
            os.remove(file_path)
            return response


        return HttpResponse('completed ...')
    
    else:
        return HttpResponse('permission Denied')
    
@login_required
def create_single_payout(request):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        current_time = localtime(timezone.now())

        last_time = current_time.replace(hour=6, minute=0, second=0, microsecond=0)
        
        all_theatres = Theatre.objects.all()
        pending_payments = theatre_payment.objects.filter(is_settled=False, status='Success', payment_method='Gateway', time__lt=last_time)

        data = []

        end_str = last_time.strftime("%d %b %Y")

        for theatre in all_theatres:
            theatre_payments = pending_payments.filter(order__seat__row__hall__theatre=theatre)
            last_payout = theatre.payoutlogs_set.last()

            try:
                if last_payout == None:
                    first_payment = theatre_payments.first()
                    start_str = first_payment.time.strftime("%d %b %Y")
                    start_time = localtime(first_payment.time)
                
                else:
                    first_payment = last_payout.end_time
                    start_time = localtime(first_payment)
                    start_str = first_payment.strftime("%d %b %Y")
                    
                message = f"{theatre.name} from {start_str} 6 AM to {end_str} 6 AM"

                start_time = start_time.replace(hour=6, minute=0, second=0, microsecond=0)

                bank_details = theatre.bankdetails
                current_balance = sum(payment.order.full_payment() - payment.settlement for payment in theatre_payments)
                if current_balance == 0:
                    pass
                else:
                    # append the dat in execl
                    append_data = [
                        bank_details.account_name,
                        bank_details.account_number,
                        bank_details.ifsc_code,
                        current_balance,
                        message
                    ]
                    data.append(append_data)

                    # create a payout for theatre.
                    payout = PayOutLogs.objects.create(start_time=start_time, end_time=last_time, user=request.user, amount=current_balance, theatre=theatre)
                    for payment in theatre_payments:
                        payment.payout = payout
                        payment.is_settled = True
                        payment.save()

            except Exception as e:
                pass
        
        file_path = os.path.join(django_settings.MEDIA_ROOT, 'single_payout.xlsx')
        if os.path.exists(file_path):
            os.remove(file_path)
        
        wb = Workbook()
        ws = wb.active

        headers = ["Beneficiary Name (Mandatory) Full name of the customer - eg: Bruce Wayne"
                , "Beneficiary Account number (Mandatory) Beneficiary Account number to which the money should be transferred",
                "IFSC code (Mandatory) IFSC code of beneficary's bank. eg:KKBK0000958",
                "Amount (Mandatory) Amount that needs to be transfered. Eg: 100.00",
                "Description / Purpose (Optional) For Internal Reference eg: For salary",
                ]
        
        ws.append(headers)

        f_name = f'{last_time.strftime("%d-%b-%Y")}:6:00-am.xlsx'
        # Append Data
        for row in data:
            ws.append(row)
        
        wb.save(file_path)
        # ✅ Return file as response
        with open(file_path, "rb") as excel:
            response = HttpResponse(excel.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = f'attachment; filename={f_name}'

        # Clean up temp file
        os.remove(file_path)
        return response


    else:
        return HttpResponse('Permission Denied')
    
@login_required
def all_queries(request):
    all_queries = Query.objects.all()[::-1]
    paginator = Paginator(all_queries, 20)  # Show 20 Theatre Per Page.

    page_number = request.GET.get('page')
    if page_number == None:
        page_number = 1
    all_queries = paginator.get_page(page_number)
    return render(request, 'adminPortal/all-queries.html', {'all_queries': all_queries})

@login_required
def update_query(request, pk):
    return_data = {}
    try:
        query = Query.objects.get(pk=pk)
        query.seen = not query.seen
        query.save()
        return_data['status'] = query.seen
    except:
        return_data['error'] ='Not able to get id'

    return JsonResponse(return_data)

@login_required
def compare_settlment_and_payout(request):
    settlement_amount = 0
    theatre_amount = 0
    return_data = {}
    try:
        last_payout = PayOutLogs.objects.last()
        last_payout_end_date = last_payout.end_time.date()
        gateway_details = PaymentGateway.objects.get(name='Razorpay')
        client = razorpay.Client(auth=(gateway_details.gateway_key, gateway_details.gateway_secret))
        
        last_settlements = client.settlement.all()
        last_settlements = last_settlements['items']
        # get the settlement amount
        # run a for loop on all the settlements
        for settlement in last_settlements:
            timestamp = settlement['created_at']
            # we get the time of settlment
            settlment_time = localtime(timezone.datetime.fromtimestamp(timestamp, tz=dt_timezone.utc))
            settlement_date = settlment_time.date()

            if settlement_date > last_payout_end_date:
                # add the amount
                settlement_amount += settlement['amount']

        # convert paise to rupee
        settlement_amount = settlement_amount / 100
        
        # get the pending payments
        current_time = localtime(timezone.now())
        last_time = current_time.replace(hour=6, minute=0, second=0, microsecond=0)
        pending_payments = theatre_payment.objects.filter(is_settled=False, status='Success', payment_method='Gateway', time__lt=last_time)
        all_theatres = Theatre.objects.all()

        for theatre in all_theatres:
            theatre_payments = pending_payments.filter(order__seat__row__hall__theatre=theatre)
            last_payout = theatre.payoutlogs_set.last()

            try:
                bank_details = theatre.bankdetails
                current_balance = sum(payment.order.full_payment() for payment in theatre_payments)
                theatre_amount += current_balance
            except:
                pass
        
        return_data['settlement_amount'] = settlement_amount
        return_data['theatre_amount'] = theatre_amount
        profit = settlement_amount - theatre_amount
        profit = round(profit, 2)
        return_data['profit'] = profit
        if profit < 0:
            message = f"🎬 *Settlement Summary*%0A%0A Today's Settlement is not come Yet, so please wait to have the settlement complete from Razorpay"
        else:
            message = f"🎬 *Settlement Summary*%0A%0A✅ Total Settlement Amount: *₹{settlement_amount}*%0A%0A🏢 Theatre Share: *₹{theatre_amount}*%0A%0A💰 Your Profit: *₹{profit}*%0A%0A Thank you!"

        update_websocket(
            message='otp',
            customer_message=message,
            customer_phone='7206662805'
        )
    except Exception as e:
        return_data['error'] =  str(e)
    
    return JsonResponse(return_data)

@login_required
def all_refund_queries(request):
    permission = Permission.objects.get(codename="view_order")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        date_range = request.GET.get('daterange', "")
        resolve_status = request.GET.get('resolve_status', "")
        selected_theatre = request.GET.get('selected-theatre', "")

        if date_range == "":
            start_date = localtime(timezone.now()).date()
            end_date = localtime(timezone.now()).date()
        else:
            try:
                start_date = date_range.split(" - ")[0]
                end_date = date_range.split(" - ")[1]

                start_date = datetime.strptime(start_date, "%d/%b/%Y").date()
                end_date = datetime.strptime(end_date, "%d/%b/%Y").date()
            except:
                start_date = localtime(timezone.now()).date()
                end_date = localtime(timezone.now()).date()

        refund_queries = OrderRefundRequest.objects.filter(time__date__range=(start_date, end_date))
        
        if selected_theatre != "":
            refund_queries = refund_queries.filter(order__seat__row__hall__theatre__pk=selected_theatre)

        refund_queries = refund_queries.order_by('-time')

        if resolve_status == "resolved":
            refund_queries = refund_queries.filter(resolve_status=True)
            
        elif resolve_status == "not-resolved":
            refund_queries = refund_queries.filter(resolve_status=False)
            
        paginator = Paginator(refund_queries, 10)  # Show 10 items per page.

        page_number = request.GET.get('page')
        if page_number == None:
            page_number = 1

        refund_queries = paginator.get_page(page_number)

        context = {
            'refund_queries': refund_queries,
            'start_date': start_date,
            'end_date': end_date,
            'resolve_status': resolve_status,
            'selected_theatre': selected_theatre,
            }
            
        return render(request, 'adminPortal/all-refund-queries.html', context)
    else:
        return HttpResponse('Permission Denied')

def download_hall_qr(request, pk):
    file_path = os.path.join(django_settings.MEDIA_ROOT, f'hall-{pk}.png')

    def create_qr(text, output_png="qr.png", px=3840, ecc="M", margin=4, dark="black", light="white"):
        # Build QR
        qr = segno.make(text, error=ecc.lower())

        # Compute crisp integer scale
        sym_w, sym_h = qr.symbol_size(border=0)
        total_modules = max(sym_w, sym_h) + 2 * margin

        scale = px // total_modules
        if scale < 1:
            raise ValueError(
                f"Requested {px}px too small for this data. Minimum required ≈ {total_modules}px."
            )

        final_size = scale * total_modules

        # Save PNG
        qr.save(
            output_png,
            kind="png",
            scale=scale,
            border=margin,
            dark=dark,
            light=light,
        )

        print(f"QR saved: {output_png} ({final_size}×{final_size}px) ECC={ecc} margin={margin}")

    create_qr(
        text=f"https://scan2food.com/theatre/hall-qr/{pk}",
        output_png=file_path
    )

    with open(file_path, "rb") as file:
            response = HttpResponse(file.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = f'attachment; filename=hall-{pk}-qr.png'

    # Clean up temp file
    os.remove(file_path)
    return response


@login_required
def get_db_files(request):
    permission = Permission.objects.get(codename="change_theatre")
    if request.user.groups.first().permissions.filter(id=permission.id).exists():
        backups = []

        media_root = django_settings.MEDIA_ROOT
        backup_folder = os.path.join(media_root, 'backup_db')

        for filename in os.listdir(backup_folder):
            file_path = os.path.join(backup_folder, filename)
            file_size = os.path.getsize(file_path)
            file_info = {
                'name': filename,
                'size': round(file_size/1048576, 2),  # size in MB
                'path': file_path,
                "modified": filename.replace('app_backup_', '').replace(".sql", ""),
            }
            backups.append(file_info)
        
        backups.sort(key=lambda x: datetime.strptime(x["modified"], "%d-%b-%Y"), reverse=True)
        
        return render(request, 'adminPortal/db-backups.html', {'backups': backups})
    else:
        return HttpResponse('Permission Denied')
    